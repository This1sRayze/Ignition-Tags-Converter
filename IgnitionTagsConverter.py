import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from pathlib import Path
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl import load_workbook
import threading
import sys
import subprocess
import os

class IgnitionTagsGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("Ignition Tags Generator")
        self.root.geometry("900x600")
        self.root.configure(bg='#f0f0f0')
        
        # Data storage
        self.tags_data = None
        self.tags2_data = None
        self.xlsx_data = None
        self.xlsx_sheets = []
        self.current_sheet_data = None
        self.column_mapping = {}
        
        # Parameter values
        self.driver_device_connection_var = tk.StringVar(value="[Siemens Enhanced]")
        self.data_block_override_var = tk.StringVar()
        self.plc_instance_var = tk.StringVar(value="PLC")
        self.server_name_var = tk.StringVar(value="Ignition OPC UA Server")
        self.namespace_var = tk.StringVar(value="1")
        
        # File path variables
        self.udt_path_var = tk.StringVar()
        self.xlsx_path_var = tk.StringVar()
        
        self.setup_styles()
        self.setup_ui()
    
    def setup_styles(self):
        """Configure custom styles"""
        style = ttk.Style()
        style.theme_use('clam')
        
        bg_color = '#f0f0f0'
        frame_bg = '#ffffff'
        accent_color = '#0078d4'
        
        style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), background=bg_color, foreground='#2c3e50')
        style.configure('Header.TLabel', font=('Segoe UI', 10, 'bold'), background=frame_bg, foreground='#2c3e50')
        style.configure('Info.TLabel', font=('Segoe UI', 9), background=frame_bg, foreground='#555555')
        style.configure('Card.TFrame', background=frame_bg, relief='flat', borderwidth=1)
        style.configure('Card.TLabelframe', background=frame_bg, relief='solid', borderwidth=1)
        style.configure('Card.TLabelframe.Label', font=('Segoe UI', 10, 'bold'), foreground='#2c3e50', background=frame_bg)
        
        # Configure combobox and entry styles
        style.configure('TCombobox', font=('Segoe UI', 9))
        style.configure('TEntry', font=('Segoe UI', 9))
        style.configure('TButton', font=('Segoe UI', 9))
    
    def setup_ui(self):
        """Setup the main UI with modern design"""
        # Outer container
        outer_container = tk.Frame(self.root, bg='#f0f0f0')
        outer_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_label = ttk.Label(outer_container, text="🚀 Ignition Tags Generator", style='Title.TLabel')
        title_label.pack(pady=(0, 20))
        
        # Main container with stacked layout (top: controls, bottom: log)
        main_container = tk.Frame(outer_container, bg='#f0f0f0')
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Top section with two columns (left: files, right: parameters)
        top_section = tk.Frame(main_container, bg='#f0f0f0')
        top_section.pack(fill=tk.X, pady=(0, 10))
        
        left_column = tk.Frame(top_section, bg='#f0f0f0')
        left_column.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        right_column = tk.Frame(top_section, bg='#f0f0f0')
        right_column.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        # ===== LEFT COLUMN =====
        
        # File Selection Card
        file_frame = ttk.LabelFrame(left_column, text="📁 File Selection", style='Card.TLabelframe', padding=15)
        file_frame.pack(fill=tk.X)
        
        # UDT JSON file
        udt_row = tk.Frame(file_frame, bg='#ffffff')
        udt_row.pack(fill=tk.X, pady=5)
        
        udt_btn = tk.Button(udt_row, text="Select UDT JSON", command=self.browse_udt_json,
                           bg='#0078d4', fg='white', font=('Segoe UI', 9, 'bold'),
                           relief='flat', padx=15, pady=5, cursor='hand2', width=12)
        udt_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.udt_label = ttk.Label(udt_row, text="No file selected", style='Info.TLabel')
        self.udt_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # XLSX file
        xlsx_row = tk.Frame(file_frame, bg='#ffffff')
        xlsx_row.pack(fill=tk.X, pady=5)
        
        xlsx_btn = tk.Button(xlsx_row, text="Select XLSX File", command=self.browse_xlsx,
                            bg='#5c6bc0', fg='white', font=('Segoe UI', 9, 'bold'),
                            relief='flat', padx=15, pady=5, cursor='hand2', width=12)
        xlsx_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.xlsx_label = ttk.Label(xlsx_row, text="No file selected", style='Info.TLabel')
        self.xlsx_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Export Button
        export_btn = tk.Button(file_frame, text="📊 Export All Sheets", command=self.export_all_sheets,
                              bg='#28a745', fg='white', font=('Segoe UI', 10, 'bold'),
                              relief='flat', padx=20, pady=8, cursor='hand2',
                              activebackground='#218838', activeforeground='white')
        export_btn.pack(fill=tk.X, padx=5, pady=10)
        
        # ===== RIGHT COLUMN =====
        
        # Tag Parameters Card
        params_frame = ttk.LabelFrame(right_column, text="🏷️ Tag Parameters", style='Card.TLabelframe', padding=15)
        params_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Driver Device Connection
        ttk.Label(params_frame, text="Driver Device Connection:", style='Info.TLabel').grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(params_frame, textvariable=self.driver_device_connection_var, width=35).grid(row=0, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        
        # PLCInstance
        ttk.Label(params_frame, text="PLCInstance:", style='Info.TLabel').grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(params_frame, textvariable=self.plc_instance_var, width=35).grid(row=1, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        
        # ServerName
        ttk.Label(params_frame, text="ServerName:", style='Info.TLabel').grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(params_frame, textvariable=self.server_name_var, width=35).grid(row=2, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        
        # Data Block Override
        ttk.Label(params_frame, text="Data Block Override:", style='Info.TLabel').grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(params_frame, textvariable=self.data_block_override_var, width=35).grid(row=3, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        
        # Namespace
        ttk.Label(params_frame, text="Namespace:", style='Info.TLabel').grid(row=4, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(params_frame, textvariable=self.namespace_var, width=35).grid(row=4, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        
        # Logs Card - Full width at bottom
        log_frame = ttk.LabelFrame(main_container, text="📝 Processing Log", style='Card.TLabelframe', padding=15)
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        log_container = tk.Frame(log_frame, bg='#ffffff')
        log_container.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = tk.Text(log_container, wrap=tk.WORD, bg='#f8f9fa',
                               fg='#2c3e50', font=('Consolas', 9), relief='flat', padx=10, pady=10)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(log_container, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
    
    def log(self, message):
        """Log message to the log text widget"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update()
    
    def browse_udt_json(self):
        file = filedialog.askopenfilename(filetypes=[("JSON files", "*.json"), ("All files", "*.*")])
        if file:
            self.udt_path_var.set(file)
            try:
                with open(file, 'r', encoding='utf-8') as f:
                    self.tags_data = json.load(f)
                
                # Extract parameters from UDT definitions
                self.extract_parameters_from_udts()
                
                self.udt_label.config(text=f"✓ {Path(file).name}")
                self.log(f"✓ Loaded UDT JSON: {Path(file).name}")
            except Exception as e:
                self.log(f"✗ Error loading UDT JSON: {str(e)}")
                messagebox.showerror("Error", f"Failed to load UDT JSON: {e}")
    
    def extract_parameters_from_udts(self):
        """Extract parameters from UDT definitions and populate UI"""
        if not self.tags_data:
            return
        
        # Look for the first UDT with parameters
        for tag_folder in self.tags_data.get('tags', []):
            if tag_folder.get('tagType') == 'Folder':
                for udt in tag_folder.get('tags', []):
                    if udt.get('tagType') == 'UdtType':
                        params = udt.get('parameters', {})
                        
                        # Extract and set parameters
                        if 'DriverDeviceConnection' in params:
                            value = params['DriverDeviceConnection'].get('value', '[Siemens Enhanced]')
                            self.driver_device_connection_var.set(str(value))
                        
                        if 'ServerName' in params:
                            value = params['ServerName'].get('value', 'Ignition OPC UA Server')
                            self.server_name_var.set(str(value))
                        
                        if 'PLCInstance' in params:
                            value = params['PLCInstance'].get('value', 'PLC')
                            self.plc_instance_var.set(str(value))
                        
                        if 'Namespace' in params:
                            value = params['Namespace'].get('value', '1')
                            self.namespace_var.set(str(value))
                        
                        # Found first UDT with parameters, exit
                        return
    
    def browse_xlsx(self):
        """Browse for XLSX file"""
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file:
            self.xlsx_path_var.set(file)
            try:
                wb = load_workbook(file, data_only=True)
                self.xlsx_sheets = wb.sheetnames
                
                self.xlsx_label.config(text=f"✓ {Path(file).name}")
                self.log(f"✓ Loaded XLSX: {Path(file).name} ({len(self.xlsx_sheets)} sheets)")
            except Exception as e:
                self.log(f"✗ Error loading XLSX: {str(e)}")
                messagebox.showerror("Error", f"Failed to load XLSX: {e}")
    
    def on_sheet_selected(self, event=None):
        """Load selected sheet data"""
        if not self.xlsx_path_var.get():
            return
        
        try:
            wb = load_workbook(self.xlsx_path_var.get(), data_only=True)
            # Get first sheet
            ws = wb[wb.sheetnames[0]]
            
            # Get headers
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value)
            
            # Load data rows
            self.current_sheet_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        row_dict[header] = row[idx]
                self.current_sheet_data.append(row_dict)
            
            self.log(f"✓ Loaded sheet with {len(self.current_sheet_data)} rows")
            self.log(f"  Columns: {', '.join(headers)}")
        except Exception as e:
            self.log(f"✗ Error loading sheet: {str(e)}")
            messagebox.showerror("Error", f"Failed to load sheet: {e}")
    
    def find_udt_definition(self, udt_type: str) -> Optional[Dict]:
        """Find UDT definition by type name"""
        if not self.tags_data:
            return None
        
        for tag_folder in self.tags_data.get('tags', []):
            if tag_folder.get('tagType') == 'Folder':
                for udt in tag_folder.get('tags', []):
                    if udt.get('name') == udt_type and udt.get('tagType') == 'UdtType':
                        return udt
        return None
    
    def find_folder_for_udt(self, udt_type: str) -> str:
        """Find the folder containing a UDT type"""
        if not self.tags_data:
            return ''
        
        for tag_folder in self.tags_data.get('tags', []):
            if tag_folder.get('tagType') == 'Folder':
                for udt in tag_folder.get('tags', []):
                    if udt.get('name') == udt_type:
                        return tag_folder.get('name', '')
        return ''
    
    def create_tags_from_udt(self, udt_tags: List[Dict]) -> List[Dict]:
        """Create tag structure from UDT definition"""
        tags = []
        for tag in udt_tags:
            new_tag = {
                "name": tag.get('name', ''),
                "tagType": tag.get('tagType', 'AtomicTag')
            }
            if 'tags' in tag and tag['tags']:
                new_tag['tags'] = self.create_tags_from_udt(tag['tags'])
            tags.append(new_tag)
        return tags
    
    def is_atomic_type(self, type_str: str) -> bool:
        """Check if type is an atomic data type (not UDT)"""
        atomic_types = [
            'INT', 'REAL', 'BOOL', 'STRING', 'DINT', 'LINT', 'FLOAT', 'DOUBLE',
            'BYTE', 'SINT', 'USINT', 'UINT', 'UDINT', 'ULINT'
        ]
        return type_str.upper() in atomic_types
    
    def create_tag_instance(self, row: Dict, udt_type: str) -> Dict:
        """Create a tag instance (either UDT or Atomic)"""
        tag_name_col = self.mapping_vars['Tag Name'].get()
        data_block_col = self.mapping_vars['Data Block'].get()
        
        tag_name = row.get(tag_name_col, '')
        data_block = row.get(data_block_col, '')
        
        # Check if this is an atomic type
        if self.is_atomic_type(udt_type):
            return self.create_atomic_tag(tag_name, udt_type, data_block, row)
        else:
            return self.create_udt_instance(tag_name, udt_type, data_block)
    
    def create_tag_instance_simple(self, tag_name: str, udt_type: str, data_block: str, tag_history: str = '') -> Dict:
        """Create a tag instance without requiring mapping (for auto-detected columns)"""
        # Check if this is an atomic type
        if self.is_atomic_type(udt_type):
            return self.create_atomic_tag(tag_name, udt_type, data_block, {})
        else:
            return self.create_udt_instance(tag_name, udt_type, data_block, tag_history)
    
    def create_atomic_tag(self, tag_name: str, data_type: str, data_block: str, row: Dict) -> Dict:
        """Create an OPC UA item tag (not memory tag) with auto-generated path"""
        # Map common data types to Ignition data types
        type_mapping = {
            'INT': 'Int2',
            'REAL': 'Float4',
            'FLOAT': 'Float4',
            'DOUBLE': 'Float8',
            'BOOL': 'Boolean',
            'STRING': 'String',
            'DINT': 'Int4',
            'LINT': 'Int8',
            'BYTE': 'Uint1',
            'SINT': 'Int1',
            'USINT': 'Uint1',
            'UINT': 'Uint2',
            'UDINT': 'Uint4',
            'ULINT': 'Uint8',
            'ANALOG': 'Float4',
            'DIGITAL': 'Boolean',
            'CALCULATED': 'Float4',
            'COMM': 'String'
        }
        
        ignition_type = type_mapping.get(data_type.upper(), 'String')
        
        # Get parameters from UI
        namespace = self.namespace_var.get()
        driver_device = self.driver_device_connection_var.get()
        plc_instance = self.plc_instance_var.get()
        server_name = self.server_name_var.get()
        data_block_value = self.data_block_override_var.get() if self.data_block_override_var.get() else data_block
        
        # Construct OPC UA tag path - format: ns=<namespace>;s=<DriverDevice><PLCInstance>.Blocks.<DataBlock>.<TagName>
        opc_ua_path = f"ns={namespace};s={driver_device}{plc_instance}.Blocks.{data_block_value}.{tag_name}"
        
        # Create OPC UA item tag with correct Ignition format
        tag_instance = {
            "dataType": ignition_type,
            "name": tag_name,
            "opcItemPath": opc_ua_path,
            "opcServer": server_name,
            "tagGroup": "Normal_Scan",
            "tagType": "AtomicTag",
            "valueSource": "opc"
        }
        
        return tag_instance
    
    def create_udt_instance(self, tag_name: str, udt_type: str, data_block: str, tag_history: str = '') -> Dict:
        """Create a UDT instance tag"""
        # Use UDTS/ prefix with just the UDT type name (no folder path)
        type_id = f"UDTS/{udt_type}"
        
        # Use override data block if provided, otherwise use the one from row
        final_data_block = self.data_block_override_var.get() if self.data_block_override_var.get() else data_block
        
        # Normalize tag history value: convert to "True" or "False" string
        historized_value = "True"
        if tag_history:
            tag_history_lower = str(tag_history).lower().strip()
            historized_value = "true" if tag_history_lower in ['true', '1', 'yes'] else "false"
        
        # Create base tag instance with all parameters
        tag_instance = {
            "name": tag_name,
            "typeId": type_id,
            "parameters": {
                "DriverDeviceConnection": {
                    "dataType": "String",
                    "value": self.driver_device_connection_var.get()
                },
                "DataBlock": {
                    "dataType": "String",
                    "value": str(final_data_block) if final_data_block else ""
                },
                "PLCInstance": {
                    "dataType": "String",
                    "value": self.plc_instance_var.get()
                },
                "ServerName": {
                    "dataType": "String",
                    "value": self.server_name_var.get()
                },
                "Namespace": {
                    "dataType": "Int4",
                    "value": self.namespace_var.get()
                },
                "Tag_Historized": {
                    "dataType": "String",
                    "value": historized_value
                }
            },
            "tagType": "UdtInstance",
            "tags": []
        }
        
        # Copy the tag structure from UDT definition if it exists
        udt_def = self.find_udt_definition(udt_type)
        if udt_def and 'tags' in udt_def:
            tag_instance['tags'] = self.create_tags_from_udt(udt_def.get('tags', []))
        
        return tag_instance
    
    def export_all_sheets(self):
        """Export JSON for all sheets (except SCADA Signal)"""
        # Validate inputs
        if not self.tags_data:
            messagebox.showerror("Error", "Please load UDT JSON first")
            return
        
        if not self.xlsx_path_var.get():
            messagebox.showerror("Error", "Please load XLSX file first")
            return
        
        # Ask for output file location BEFORE starting
        xlsx_path = self.xlsx_path_var.get()
        default_filename = f"{Path(xlsx_path).stem}_all_sheets_ignition.json"
        output_file = filedialog.asksaveasfilename(
            title="Save JSON File As",
            initialdir=str(Path(xlsx_path).parent),
            initialfile=default_filename,
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if not output_file:
            self.log("✗ Output file selection cancelled")

        self.output_file = output_file
        
        # Run export in background thread
        thread = threading.Thread(target=self._export_all_sheets_thread)
        thread.start()
    
    def _export_all_sheets_thread(self):
        """Export all sheets to JSON in background thread"""
        try:
            self.log("\n" + "="*70)
            self.log("🚀 Starting export all sheets...")
            self.root.update()
            
            # Get column names from XLSX headers (auto-detect)
            wb_temp = load_workbook(self.xlsx_path_var.get(), data_only=True)
            first_sheet = wb_temp[wb_temp.sheetnames[0]]
            
            headers = []
            for cell in first_sheet[1]:
                if cell.value:
                    headers.append(cell.value)
            
            # Auto-detect column positions
            tag_name_col = None
            data_block_col = None
            udt_type_col = None
            tag_history_col = None
            
            for header in headers:
                if 'tag' in header.lower() and 'name' in header.lower():
                    tag_name_col = header
                if 'data' in header.lower() and 'block' in header.lower():
                    data_block_col = header
                if 'udt' in header.lower() and 'type' in header.lower():
                    udt_type_col = header
                if 'tag' in header.lower() and 'history' in header.lower():
                    tag_history_col = header
            
            if not all([tag_name_col, data_block_col, udt_type_col]):
                self.log(f"⚠️  Could not auto-detect all required columns")
                self.log(f"  Found: Tag Name={tag_name_col}, Data Block={data_block_col}, UDT Type={udt_type_col}")
                messagebox.showerror("Error", "Could not auto-detect required columns in XLSX")
                return
            
            ignored_sheets = ["SCADA Signal", "scada signal","scada_signal","SCADA_SIGNAL"]
            
            # Load workbook
            wb = load_workbook(self.xlsx_path_var.get(), data_only=True)
            sheet_names = wb.sheetnames
            
            self.log(f"📊 Found {len(sheet_names)} sheets, processing...")
            
            # Create output structure for all sheets - organized by tab -> data block
            all_sheet_folders = {}
            total_udt = 0
            total_atomic = 0
            total_skipped = 0
            processed_sheets = 0
            
            for sheet_name in sheet_names:
                # Skip SCADA Signal sheets
                if any(sheet_name.lower() == ignored.lower() for ignored in ignored_sheets):
                    self.log(f"⏭️  Skipping sheet: {sheet_name}")
                    continue
                
                try:
                    ws = wb[sheet_name]
                    
                    # Load data rows
                    sheet_data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {}
                        for idx, header in enumerate(headers):
                            if idx < len(row):
                                row_dict[header] = row[idx]
                        sheet_data.append(row_dict)
                    
                    self.log(f"📄 Processing '{sheet_name}' ({len(sheet_data)} rows)...")
                    
                    # Process each row for this sheet
                    processed_tags = set()
                    sheet_udt = 0
                    sheet_atomic = 0
                    
                    for row in sheet_data:
                        tag_name = row.get(tag_name_col, '')
                        udt_type = row.get(udt_type_col, '')
                        data_block = row.get(data_block_col, '') if data_block_col else ''
                        tag_history = row.get(tag_history_col, '') if tag_history_col else ''
                        
                        # Skip if already processed or empty
                        tag_key = f"{sheet_name}_{tag_name}_{udt_type}"
                        if tag_key in processed_tags:
                            self.log(f"  ⏭️  Skipped '{tag_name}' - duplicate entry")
                            total_skipped += 1
                            continue
                        if not tag_name:
                            self.log(f"  ⏭️  Skipped - empty tag name")
                            total_skipped += 1
                            continue
                        if not udt_type:
                            self.log(f"  ⏭️  Skipped '{tag_name}' - empty UDT type")
                            total_skipped += 1
                            continue
                        
                        processed_tags.add(tag_key)
                        
                        # Create tag instance
                        tag_instance = self.create_tag_instance_simple(tag_name, udt_type, data_block, tag_history)
                        
                        # Organize by sheet name (tab) -> data block as nested folders
                        sheet_folder_name = sheet_name if sheet_name else "Ungrouped"
                        db_folder_name = str(data_block) if data_block else "NoDataBlock"
                        
                        # Create sheet folder if not exists
                        if sheet_folder_name not in all_sheet_folders:
                            all_sheet_folders[sheet_folder_name] = {
                                "name": sheet_folder_name,
                                "tagType": "Folder",
                                "tags": []
                            }
                        
                        # Find or create data block folder within sheet folder
                        db_folder = None
                        for tag in all_sheet_folders[sheet_folder_name]['tags']:
                            if tag.get('name') == db_folder_name and tag.get('tagType') == 'Folder':
                                db_folder = tag
                                break
                        
                        if db_folder is None:
                            db_folder = {
                                "name": db_folder_name,
                                "tagType": "Folder",
                                "tags": []
                            }
                            all_sheet_folders[sheet_folder_name]['tags'].append(db_folder)
                        
                        # Add tag to data block folder
                        db_folder['tags'].append(tag_instance)
                        
                        if self.is_atomic_type(udt_type):
                            total_atomic += 1
                            sheet_atomic += 1
                        else:
                            total_udt += 1
                            sheet_udt += 1
                    
                    self.log(f"  ✓ {sheet_udt} UDT + {sheet_atomic} atomic tags")
                    processed_sheets += 1
                
                except Exception as sheet_error:
                    self.log(f"⚠️  Error processing sheet '{sheet_name}': {str(sheet_error)}")
                    continue
            
            output_path = Path(self.output_file)
            
            # Create output with all folders
            output = {"tags": list(all_sheet_folders.values())}
            
            # Save output
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(output, f, indent=2, ensure_ascii=False)
            
            total = total_udt + total_atomic
            status_msg = f"✓ Exported {processed_sheets} sheets with {total} tags: {total_udt} UDT + {total_atomic} atomic"
            if total_skipped > 0:
                status_msg += f" (Skipped {total_skipped} duplicates)"
            
            self.log(f"\n{status_msg}")
            self.log(f"💾 Saved to: {output_path.name}")
            self.log("="*70)
            
            messagebox.showinfo("Success", f"{status_msg}\n\nSaved to:\n{output_path}")
            
            # Open the output folder
            output_folder = str(output_path.parent)
            if sys.platform == 'win32':
                os.startfile(output_folder)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', output_folder])
            else:
                subprocess.Popen(['xdg-open', output_folder])
        
        except Exception as e:
            self.log(f"\n✗ Error: {str(e)}")
            messagebox.showerror("Error", f"Failed to export all sheets: {e}")

def main():
    root = tk.Tk()
    app = IgnitionTagsGenerator(root)
    root.mainloop()

if __name__ == "__main__":
    main()
